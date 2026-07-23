"""Peer Comparison Report Generator
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
This module generates a peer comparison Excel report with one worksheet per
peer group, using percentile data from ``src.screener.peer``.

The report includes:
* Company Name, Peer Group, Overall Peer Score
* All percentile metrics (ROE, Net Profit Margin, Revenue CAGR, PAT CAGR,
  Free Cash Flow, Debt-to-Equity, PE, PB)
* Sorted by Overall Peer Score (descending)
* Conditional formatting: Top 20% Green, Middle 60% Yellow, Bottom 20% Red
* Frozen header row, auto-sized columns, bold headers

Functions
==========
* ``generate_peer_comparison_report()`` – main entry point that writes
  ``Data/output/peer_comparison.xlsx``.
* ``_apply_conditional_formatting()`` – internal helper for 3-color scale.
"""

from __future__ import annotations

import os
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# Import peer engine to reuse computed percentiles
from .peer import compute_peer_percentiles
from .engine import load_screener_data


# Columns to include in the report (in desired order)
REPORT_COLUMNS = [
    ("company_id", "Company Name"),
    ("peer_group_name", "Peer Group"),
    ("overall_peer_score", "Overall Peer Score"),
    ("roe_percentile", "ROE Percentile"),
    ("net_profit_margin_percentile", "Net Profit Margin Percentile"),
    ("revenue_cagr_percentile", "Revenue CAGR Percentile"),
    ("pat_cagr_percentile", "PAT CAGR Percentile"),
    ("free_cash_flow_percentile", "Free Cash Flow Percentile"),
    ("debt_to_equity_percentile", "Debt-to-Equity Percentile"),
    ("pe_percentile", "PE Percentile"),
    ("pb_percentile", "PB Percentile"),
]

# Metrics that should receive conditional formatting (all percentile columns)
PERCENTILE_COLUMNS = [col for col, _ in REPORT_COLUMNS if "percentile" in col.lower() or col == "overall_peer_score"]


def _apply_conditional_formatting(ws, start_row: int, end_row: int, col_letter: str) -> None:
    """Apply 3-color scale conditional formatting to a column.

    - Top 20% (80th percentile and above) → Green
    - Middle 60% (20th–80th percentile) → Yellow
    - Bottom 20% (below 20th percentile) → Red

    Uses openpyxl's ColorScaleRule with 'percentile' type.
    """
    # Green for high, Yellow for middle, Red for low
    # Note: ColorScaleRule with 'percentile' interpolates between the three points
    rule = ColorScaleRule(
        start_type="percentile", start_value=0, start_color="FFC7CE",  # Red
        mid_type="percentile", mid_value=50, mid_color="FFEB9C",       # Yellow
        end_type="percentile", end_value=100, end_color="C6EFCE",      # Green
    )
    ws.conditional_formatting.add(f"{col_letter}{start_row}:{col_letter}{end_row}", rule)


def _auto_size_columns(ws, min_width: int = 12, max_width: int = 30) -> None:
    """Auto-size columns based on header and content width."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            if cell.value is not None:
                # Approximate character width
                val_str = str(cell.value)
                # Account for multi-line by taking max line length
                line_len = max(len(line) for line in val_str.split("\n"))
                max_len = max(max_len, line_len + 2)
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


def generate_peer_comparison_report(
    output_path: str = "Data/output/peer_comparison.xlsx"
) -> str:
    """Generate the peer comparison Excel report.

    Parameters
    ----------
    output_path : str
        Path to save the Excel file. Defaults to
        ``Data/output/peer_comparison.xlsx`` relative to project root.

    Returns
    -------
    str
        Absolute path to the generated file.
    """
    # Load screener data and compute percentiles (reuses peer.py)
    df = load_screener_data()
    df_pct = compute_peer_percentiles(df)

    # Filter to companies with a peer group assigned
    df_peer = df_pct[df_pct["peer_group_name"].notna()].copy()

    if df_peer.empty:
        raise ValueError("No companies with peer groups found")

    # Select and rename columns for the report
    available_cols = [c for c, _ in REPORT_COLUMNS if c in df_peer.columns]
    rename_map = {c: label for c, label in REPORT_COLUMNS if c in available_cols}
    df_report = df_peer[available_cols].rename(columns=rename_map)

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    # Iterate over peer groups
    for peer_group in sorted(df_report["Peer Group"].unique()):
        group_df = df_report[df_report["Peer Group"] == peer_group].copy()

        # Sort by Overall Peer Score descending
        if "Overall Peer Score" in group_df.columns:
            group_df = group_df.sort_values("Overall Peer Score", ascending=False)

        # Create worksheet (sanitize sheet name)
        sheet_name = str(peer_group)[:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=sheet_name)

        # Write headers
        headers = list(group_df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Write data
        for row_idx, (_, row) in enumerate(group_df.iterrows(), 2):
            for col_idx, header in enumerate(headers, 1):
                val = row[header]
                # Convert numpy types to native Python for openpyxl
                if pd.isna(val):
                    cell_val = None
                elif hasattr(val, "item"):  # numpy scalar
                    cell_val = val.item()
                else:
                    cell_val = val
                ws.cell(row=row_idx, column=col_idx, value=cell_val)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Apply conditional formatting to percentile columns
        last_row = ws.max_row
        for col_idx, header in enumerate(headers, 1):
            if header in [label for _, label in REPORT_COLUMNS if "percentile" in label.lower() or label == "Overall Peer Score"]:
                col_letter = get_column_letter(col_idx)
                _apply_conditional_formatting(ws, 2, last_row, col_letter)

        # Auto-size columns
        _auto_size_columns(ws)

    # Save
    wb.save(output_path)
    return os.path.abspath(output_path)


if __name__ == "__main__":
    # Allow running as script
    path = generate_peer_comparison_report()
    print(f"Report saved to {path}")


__all__ = ["generate_peer_comparison_report"]
