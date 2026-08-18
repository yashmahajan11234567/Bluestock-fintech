import pandas as pd
import numpy as np
from typing import Dict, Any, Optional
import os
import warnings


def load_screener_data() -> pd.DataFrame:
    """
    Load and merge the screener datasets from the raw data directory.
    Returns a DataFrame containing the most recent year's data for each company.
    """
    base_path = os.path.join(os.path.dirname(__file__), '..', '..', 'Data', 'raw')
    # Companies
    companies = pd.read_excel(os.path.join(base_path, 'companies.xlsx'))
    companies = companies.rename(columns={'id': 'company_id'})
    # Sectors
    sectors = pd.read_excel(os.path.join(base_path, 'sectors.xlsx'))
    # Financial ratios
    fin_ratio = pd.read_excel(os.path.join(base_path, 'financial_ratios.xlsx'))
    def _parse_year_to_int(y):
        if isinstance(y, str):
            import re
            m = re.search(r'\b(\d{4})\b', y)
            return int(m.group(1)) if m else 0
        elif isinstance(y, (int, float, np.integer)):
            return int(y)
        return 0
    fin_ratio['_year_int'] = fin_ratio['year'].apply(_parse_year_to_int)
    fin_ratio = fin_ratio.sort_values(['company_id', '_year_int'], ascending=[True, False])
    fin_ratio = fin_ratio.drop_duplicates(subset=['company_id'], keep='first')
    fin_ratio = fin_ratio.drop(columns=['_year_int'])
    # Market cap
    market_cap = pd.read_excel(os.path.join(base_path, 'market_cap.xlsx'))
    market_cap = market_cap.sort_values(['company_id', 'year'], ascending=[True, False])
    market_cap = market_cap.drop_duplicates(subset=['company_id'], keep='first')
    # Profit and loss
    profitandloss = pd.read_excel(os.path.join(base_path, 'profitandloss.xlsx'), header=1)
    if 'id' in profitandloss.columns:
        profitandloss = profitandloss.drop(columns=['id'])
    # Compute CAGR from profitandloss history (all years)
    # Create/normalize _year_int FIRST, then keep a copy for CAGR calculation
    if 'year' in profitandloss.columns:
        profitandloss['_year_int'] = profitandloss['year'].apply(_parse_year_to_int)
        profitandloss = profitandloss.sort_values(['company_id', '_year_int'], ascending=[True, False])
        profitandloss = profitandloss.drop_duplicates(subset=['company_id'], keep='first')
        # Now fix: sort ascending, then keep='last' gets the latest year.
        profitandloss = profitandloss.sort_values(['company_id', '_year_int'], ascending=[True, False])
        profitandloss = profitandloss.drop_duplicates(subset=['company_id'], keep='last')
        profitandloss = profitandloss.drop(columns=['_year_int'])
    else:
        profitandloss = profitandloss.drop_duplicates(subset=['company_id'], keep='last')

    # Create pl_full AFTER _year_int is available for CAGR calculation
    # Reload and prepare with _year_int for CAGR
    pl_full = pd.read_excel(os.path.join(base_path, 'profitandloss.xlsx'), header=1)
    if 'id' in pl_full.columns:
        pl_full = pl_full.drop(columns=['id'])
    if 'year' in pl_full.columns:
        pl_full['_year_int'] = pl_full['year'].apply(_parse_year_to_int)
        # Keep only valid years for CAGR (exclude TTM)
        pl_full = pl_full[pl_full['_year_int'] > 0].copy()
    # Helper functions
    def _to_float(value):
        try:
            v = pd.to_numeric(value, errors='coerce')
            return v if not pd.isna(v) else None
        except Exception:
            return None
    def _calculate_cagr_value(start_val, end_val, n_years):
        if start_val is None or end_val is None or n_years <= 0:
            return None
        if start_val == 0:
            return None
        return ((end_val / start_val) ** (1 / n_years) - 1) * 100.0
    def _calculate_cagr(group):
        # group is a DataFrame for one company with columns: year, sales, net_profit, _year_int
        # Exclude rows with invalid year (year_int <= 0) for CAGR calculation (e.g., TTM)
        group = group[group['_year_int'] > 0].copy()
        if len(group) < 2:
            return pd.Series({'compounded_sales_growth': None, 'compounded_profit_growth': None})
        # Sort by year_int for consistent ordering
        group = group.sort_values('_year_int')
        years = group['_year_int'].tolist()
        sales = group['sales'].apply(_to_float).tolist()
        net_profits = group['net_profit'].apply(_to_float).tolist()
        # Build lists of (year, value) for sales and net_profit where value is not None and > 0
        sales_pairs = [(y, s) for y, s in zip(years, sales) if s is not None and s > 0]
        net_profit_pairs = [(y, s) for y, s in zip(years, net_profits) if s is not None and s > 0]
        sales_cagr = None
        if len(sales_pairs) >= 2:
            start_year, start_val = sales_pairs[0]
            end_year, end_val = sales_pairs[-1]
            n_years = end_year - start_year
            if n_years > 0:
                sales_cagr = _calculate_cagr_value(start_val, end_val, n_years)
        net_profit_cagr = None
        if len(net_profit_pairs) >= 2:
            start_year, start_val = net_profit_pairs[0]
            end_year, end_val = net_profit_pairs[-1]
            n_years = end_year - start_year
            if n_years > 0:
                net_profit_cagr = _calculate_cagr_value(start_val, end_val, n_years)
        return pd.Series({
            'compounded_sales_growth': sales_cagr,
            'compounded_profit_growth': net_profit_cagr
        })
    # We need to apply this to each company_id group
    cagr_df = pl_full.groupby('company_id').apply(_calculate_cagr, include_groups=False).reset_index()
    # Now, we want the latest year's profitandloss data for merging (for sales, net_profit, etc.)
    # profitandloss already has the latest year's data (because we sorted and dropped duplicates keeping first? Wait, we did keep='first' after sorting by year ascending? Actually, we sorted by year ascending and then dropped duplicates keeping first -> that gives the earliest year. We want the latest year.
    # Let's recompute the latest year's profitandloss data:
    pl_latest = profitandloss.copy()
    if 'year' in pl_latest.columns:
        pl_latest['_year_int'] = pl_latest['year'].apply(_parse_year_to_int)
        pl_latest = pl_latest.sort_values(['company_id', '_year_int'], ascending=[True, False])
        pl_latest = pl_latest.drop_duplicates(subset=['company_id'], keep='first')  # keep the latest because we sorted ascending? Wait, we sorted ascending, so the latest year is at the end. We want to keep the last duplicate. So we should do keep='last'.
        # Let's fix: sort ascending, then keep='last' gets the latest year.
        pl_latest = pl_latest.sort_values(['company_id', '_year_int'], ascending=[True, False])
        pl_latest = pl_latest.drop_duplicates(subset=['company_id'], keep='last')
        pl_latest = pl_latest.drop(columns=['_year_int'])
    else:
        pl_latest = pl_latest.drop_duplicates(subset=['company_id'], keep='last')
    # Merge
    df = pd.merge(companies, sectors, on='company_id', how='left')
    df = pd.merge(df, fin_ratio, on='company_id', how='left')
    df = pd.merge(df, market_cap, on='company_id', how='left')
    # Merge latest profitandloss (for columns like sales, net_profit, etc.)
    df = pd.merge(df, pl_latest, on='company_id', how='left')
    # Merge CAGR data (overwrite the CAGR columns if they exist from profitandloss? They don't, but we will have the columns from cagr_df)
    df = pd.merge(df, cagr_df, on='company_id', how='left')
    return df


def apply_filters(df: pd.DataFrame, filters: Dict[str, Any]) -> pd.DataFrame:
    """
    Apply filters to the screener DataFrame.
    """
    filtered_df = df.copy()
    filter_configs = {
        'ROE': ('return_on_equity_pct', 'min', None, None),
        'Free Cash Flow': ('free_cash_flow_cr', 'min', None, None),
        'Operating Profit Margin': ('operating_profit_margin_pct', 'min', None, None),
        'Dividend Yield': ('dividend_yield_pct', 'min', None, None),
        'Interest Coverage': ('interest_coverage', 'min', None, 'interest_coverage'),
        'Market Cap': ('market_cap_crore', 'min', None, None),
        'Asset Turnover': ('asset_turnover', 'min', None, None),
        'Debt to Equity': ('debt_to_equity', None, 'max', 'debt_to_equity_financials'),
        'PE': ('pe_ratio', None, 'max', None),
        'PB': ('pb_ratio', None, 'max', None),
        # Additional mappings
        'Revenue CAGR': ('compounded_sales_growth', 'min', None, None),
        'PAT CAGR': ('compounded_profit_growth', 'min', None, None),
        'Dividend Payout': ('dividend_payout', None, 'max', None),
        'Sales': ('sales', 'min', None, None),
    }
    for name, value in filters.items():
        if name not in filter_configs:
            warnings.warn(f'Unknown filter "{name}". Ignoring.', UserWarning)
            continue
        col, min_key, max_key, special = filter_configs[name]
        if col not in filtered_df.columns:
            warnings.warn(f'Column "{col}" for filter "{name}" not found.', UserWarning)
            continue
        if min_key and min_key in value:
            min_val = value[min_key]
            if not pd.isna(min_val):
                if special == 'interest_coverage':
                    mask = filtered_df[col].apply(
                        lambda x: True if isinstance(x, str) and x.strip().lower() == 'debt free'
                        else pd.to_numeric(x, errors='coerce') >= min_val if not pd.isna(x) else False
                    )
                else:
                    mask = pd.to_numeric(filtered_df[col], errors='coerce') >= min_val
                filtered_df = filtered_df[mask]
        if max_key and max_key in value:
            max_val = value[max_key]
            if not pd.isna(max_val):
                if special == 'debt_to_equity_financials':
                    sector_col = 'broad_sector'
                    if sector_col not in filtered_df.columns:
                        warnings.warn('Sector column missing, applying filter to all.', UserWarning)
                        mask = pd.to_numeric(filtered_df[col], errors='coerce') <= max_val
                    else:
                        is_fin = filtered_df[sector_col].str.contains('Financial', case=False, na=False)
                        mask = is_fin | (pd.to_numeric(filtered_df[col], errors='coerce') <= max_val)
                else:
                    mask = pd.to_numeric(filtered_df[col], errors='coerce') <= max_val
                filtered_df = filtered_df[mask]
    return filtered_df


def _winsorize_and_scale(series, higher_is_better=True):
    s = pd.to_numeric(series, errors='coerce')
    nan = s.isna()
    valid = s[~nan]
    if valid.empty:
        return pd.Series(np.nan, index=series.index)
    p10, p90 = valid.quantile([0.10, 0.90])
    s = s.clip(lower=p10, upper=p90)
    mn, mx = s.min(), s.max()
    if mx == mn:
        res = pd.Series(50.0, index=series.index)
        res[nan] = np.nan
    else:
        res = (s - mn) / (mx - mn) * 100.0
        res[nan] = np.nan
    if not higher_is_better:
        res = 100.0 - res
    return res


def run_screener(filters: Optional[Dict[str, Any]] = None,
                 sort_by: str = 'return_on_equity_pct',
                 ascending: bool = False) -> pd.DataFrame:
    if filters is None:
        filters = {}
    df = load_screener_data()
    filtered_df = apply_filters(df, filters)

    # Handle edge case of empty dataframe
    if len(filtered_df) == 0:
        # Return empty dataframe with expected columns
        return filtered_df

    # Scores
    roe_score = _winsorize_and_scale(filtered_df['return_on_equity_pct'], True)
    npm_score = _winsorize_and_scale(filtered_df['net_profit_margin_pct'], True)
    profitability_score = 0.6 * roe_score + 0.4 * npm_score
    fcf_score = _winsorize_and_scale(filtered_df['free_cash_flow_cr'], True)
    pat = filtered_df['net_profit']
    cfo = filtered_df['cash_from_operations_cr']
    cfo_pat = cfo / pat.replace(0, np.nan)
    cfo_pat = cfo_pat.replace([np.inf, -np.inf], np.nan)
    cfo_pat_score = _winsorize_and_scale(cfo_pat, True)
    fcf_positive = (filtered_df['free_cash_flow_cr'] > 0).astype(float) * 100.0
    cash_quality_score = 0.5 * fcf_score + (1/3) * cfo_pat_score + (1/6) * fcf_positive
    revenue_cagr_score = _winsorize_and_scale(filtered_df['compounded_sales_growth'], True)
    pat_cagr_score = _winsorize_and_scale(filtered_df['compounded_profit_growth'], True)
    growth_score = 0.5 * revenue_cagr_score + 0.5 * pat_cagr_score
    de_score = _winsorize_and_scale(filtered_df['debt_to_equity'], False)
    ic_score = _winsorize_and_scale(filtered_df['interest_coverage'], True)
    leverage_score = (2/3) * de_score + (1/3) * ic_score
    composite_score = (0.35 * profitability_score + 0.30 * cash_quality_score +
                       0.20 * growth_score + 0.15 * leverage_score)
    filtered_df['composite_quality_score'] = composite_score

    # Sector relative score - handle edge cases
    if len(filtered_df) > 0 and 'broad_sector' in filtered_df.columns:
        try:
            def _compute_sector_score(group):
                """Compute composite sector-relative score for a single sector group."""
                roe_score = _winsorize_and_scale(group['return_on_equity_pct'], True)
                npm_score = _winsorize_and_scale(group['net_profit_margin_pct'], True)
                profitability = 0.6 * roe_score + 0.4 * npm_score

                fcf_score = _winsorize_and_scale(group['free_cash_flow_cr'], True)
                cfo_pat = group['cash_from_operations_cr'] / group['net_profit'].replace(0, np.nan)
                cfo_pat = cfo_pat.replace([np.inf, -np.inf], np.nan)
                cfo_pat_score = _winsorize_and_scale(cfo_pat, True)
                fcf_positive = (group['free_cash_flow_cr'] > 0).astype(float) * 100.0
                cash_quality = 0.5 * fcf_score + (1/3) * cfo_pat_score + (1/6) * fcf_positive

                revenue_cagr_score = _winsorize_and_scale(group['compounded_sales_growth'], True)
                pat_cagr_score = _winsorize_and_scale(group['compounded_profit_growth'], True)
                growth = 0.5 * revenue_cagr_score + 0.5 * pat_cagr_score

                de_score = _winsorize_and_scale(group['debt_to_equity'], False)
                ic_score = _winsorize_and_scale(group['interest_coverage'], True)
                leverage = (2/3) * de_score + (1/3) * ic_score

                composite = (0.35 * profitability + 0.30 * cash_quality +
                           0.20 * growth + 0.15 * leverage)
                return composite

            sector_scores = filtered_df.groupby('broad_sector', group_keys=True).apply(
                _compute_sector_score, include_groups=False
            )
            # Handle both DataFrame (single group) and Series (multiple groups) return types
            if len(sector_scores) > 0:
                if isinstance(sector_scores, pd.DataFrame):
                    # Single sector: sector_scores is a DataFrame with columns = original indices
                    # Convert to Series aligned with filtered_df index
                    sector_scores = sector_scores.iloc[0]  # Get first (only) row
                else:
                    # Multiple sectors: sector_scores is a Series with MultiIndex (sector, original_index)
                    # Drop sector level to get original index
                    sector_scores = sector_scores.droplevel(0)
                # Align with filtered_df index
                filtered_df['sector_relative_score'] = sector_scores.reindex(filtered_df.index)
        except Exception:
            # If sector scoring fails, continue without it
            pass

    if sort_by in filtered_df.columns:
        filtered_df = filtered_df.sort_values(by=sort_by, ascending=ascending, na_position='last')
    else:
        filtered_df = filtered_df.sort_values(by='return_on_equity_pct', ascending=ascending, na_position='last')
    return filtered_df


def _parse_cagr_strings(df: pd.DataFrame) -> pd.DataFrame:
    """Parse CAGR string columns (e.g. '10 Years: 21%') into numeric floats.

    This helper can be reused by other modules (e.g., peer.py) that need the
    same parsing logic without duplicating the regex code.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame containing the raw analysis columns.

    Returns
    -------
    pd.DataFrame
        A copy with ``compounded_sales_growth`` and
        ``compounded_profit_growth`` converted to float percentages.
    """
    df = df.copy()
    for col in ["compounded_sales_growth", "compounded_profit_growth"]:
        if col in df.columns:
            df[col] = (
                df[col].astype(str)
                .str.extract(r":\s*(\d+\.?\d*)\s*%")[0]
                .astype(float)
            )
    return df


def __getattr__(name):
    if name in ('load_screener_data', 'apply_filters', 'run_screener'):
        return globals()[name]
    raise AttributeError(f'module "{__name__}" has no attribute "{name}"')


# Preset filter functions
def get_quality_compounder_filters():
    return {'ROE': {'min': 15},
            'Debt to Equity': {'max': 1},
            'Free Cash Flow': {'min': 0},
            'Revenue CAGR': {'min': 10}}


def get_value_pick_filters():
    return {'PE': {'max': 20},
            'PB': {'max': 3},
            'Debt to Equity': {'max': 2},
            'Dividend Yield': {'min': 1}}


def get_growth_accelerator_filters():
    return {'PAT CAGR': {'min': 20},
            'Revenue CAGR': {'min': 15},
            'Debt to Equity': {'max': 2}}


def get_dividend_champion_filters():
    return {'Dividend Yield': {'min': 2},
            'Dividend Payout': {'max': 80},
            'Free Cash Flow': {'min': 0}}


def get_debt_free_blue_chip_filters():
    return {'Debt to Equity': {'max': 0},
            'ROE': {'min': 12},
            'Sales': {'min': 5000}}


def get_turnaround_watch_filters():
    return {'Revenue CAGR': {'min': 10},
            'Free Cash Flow': {'min': 0}}


# Output generation with conditional formatting
def generate_screener_output(output_path: str = 'Data/output/screener_output.xlsx'):
    import pandas as pd
    from openpyxl.styles import PatternFill
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    presets = {
        'Quality_Compounder': get_quality_compounder_filters,
        'Value_Pick': get_value_pick_filters,
        'Growth_Accelerator': get_growth_accelerator_filters,
        'Dividend_Champion': get_dividend_champion_filters,
        'Debt_Free_Blue_Chip': get_debt_free_blue_chip_filters,
        'Turnaround_Watch': get_turnaround_watch_filters,
    }
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet, func in presets.items():
            filt = func()
            df = run_screener(filters=filt, sort_by='composite_quality_score', ascending=False)
            cols = ['company_id','composite_quality_score','sector_relative_score',
                    'return_on_equity_pct','debt_to_equity','free_cash_flow_cr',
                    'compounded_sales_growth','dividend_yield_pct',
                    'net_profit_margin_pct','pe_ratio','pb_ratio']
            cols = [c for c in cols if c in df.columns]
            df[cols].to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.sheets[sheet]
            green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.fill = green
    print(f"Screener output saved to {output_path}")

def _winsorize_and_scale(series, higher_is_better=True):
    """Winsorize at P10/P90 and scale to [0, 100]."""
    import pandas as pd
    import numpy as np
    s = pd.to_numeric(series, errors="coerce")
    nan_mask = s.isna()
    valid = s[~nan_mask]
    if len(valid) == 0:
        return pd.Series(np.nan, index=series.index)
    p10 = valid.quantile(0.10)
    p90 = valid.quantile(0.90)
    s_w = s.clip(lower=p10, upper=p90)
    mn = s_w.min()
    mx = s_w.max()
    if mx == mn:
        r = pd.Series(np.nan, index=series.index)
        r[~nan_mask] = 50.0
    else:
        r = (s_w - mn) / (mx - mn) * 100.0
        r[nan_mask] = np.nan
    if not higher_is_better:
        r = 100.0 - r
    return r


def compute_profitability_score(df):
    """Profitability score from ROE and Operating Profit Margin."""
    import numpy as np
    roe = _winsorize_and_scale(df["return_on_equity_pct"], higher_is_better=True)
    opm = _winsorize_and_scale(df["operating_profit_margin_pct"], higher_is_better=True)
    return (roe + opm) / 2.0


def compute_cash_quality_score(df):
    """Cash quality score from Free Cash Flow and Operating Cash Flow."""
    import numpy as np
    cols = [c for c in ["free_cash_flow_cr", "cash_from_operations_cr"] if c in df.columns]
    if not cols:
        import pandas as pd
        return pd.Series(50.0, index=df.index)
    scaled = [_winsorize_and_scale(df[c], higher_is_better=True) for c in cols]
    if len(scaled) == 1:
        return scaled[0]
    return (scaled[0] + scaled[1]) / 2.0


def compute_growth_score(df):
    """Growth score placeholder: returns 50 (no CAGR data)."""
    import pandas as pd
    return pd.Series(50.0, index=df.index)


def compute_leverage_score(df):
    """Leverage score from Debt-to-Equity and Interest Coverage."""
    import numpy as np
    import pandas as pd
    # Interest Coverage: treat 'Debt Free' as max value
    ic = df["interest_coverage"]
    ic_num = pd.to_numeric(ic, errors="coerce")
    debt_free = ic.apply(lambda x: isinstance(x, str) and x.strip().lower() == "debt free")
    if debt_free.any():
        mx = ic_num.max()
        if pd.notna(mx):
            ic_num = ic_num.copy()
            ic_num[debt_free] = mx
    de = _winsorize_and_scale(df["debt_to_equity"], higher_is_better=False)
    ic_s = _winsorize_and_scale(ic_num, higher_is_better=True)
    return (de + ic_s) / 2.0



def _winsorize_and_scale(series, higher_is_better=True):
    """Winsorize at P10/P90 and scale to [0, 100]."""
    import pandas as pd
    import numpy as np
    s = pd.to_numeric(series, errors="coerce")
    nan_mask = s.isna()
    valid = s[~nan_mask]
    if len(valid) == 0:
        return pd.Series(np.nan, index=series.index)
    p10 = valid.quantile(0.10)
    p90 = valid.quantile(0.90)
    s_w = s.clip(lower=p10, upper=p90)
    mn = s_w.min()
    mx = s_w.max()
    if mx == mn:
        r = pd.Series(np.nan, index=series.index)
        r[~nan_mask] = 50.0
    else:
        r = (s_w - mn) / (mx - mn) * 100.0
        r[nan_mask] = np.nan
    if not higher_is_better:
        r = 100.0 - r
    return r


def compute_profitability_score(df):
    """Profitability score from ROE and Operating Profit Margin."""
    import numpy as np
    roe = _winsorize_and_scale(df["return_on_equity_pct"], higher_is_better=True)
    opm = _winsorize_and_scale(df["operating_profit_margin_pct"], higher_is_better=True)
    return (roe + opm) / 2.0


def compute_cash_quality_score(df):
    """Cash quality score from Free Cash Flow and Operating Cash Flow."""
    import numpy as np
    cols = [c for c in ["free_cash_flow_cr", "cash_from_operations_cr"] if c in df.columns]
    if not cols:
        import pandas as pd
        return pd.Series(50.0, index=df.index)
    scaled = [_winsorize_and_scale(df[c], higher_is_better=True) for c in cols]
    if len(scaled) == 1:
        return scaled[0]
    return (scaled[0] + scaled[1]) / 2.0


def compute_growth_score(df):
    """Growth score placeholder: returns 50 (no CAGR data)."""
    import pandas as pd
    return pd.Series(50.0, index=df.index)


def compute_leverage_score(df):
    """Leverage score from Debt-to-Equity and Interest Coverage."""
    import numpy as np
    import pandas as pd
    # Interest Coverage: treat 'Debt Free' as max value
    ic = df["interest_coverage"]
    ic_num = pd.to_numeric(ic, errors="coerce")
    debt_free = ic.apply(lambda x: isinstance(x, str) and x.strip().lower() == "debt free")
    if debt_free.any():
        mx = ic_num.max()
        if pd.notna(mx):
            ic_num = ic_num.copy()
            ic_num[debt_free] = mx
    de = _winsorize_and_scale(df["debt_to_equity"], higher_is_better=False)
    ic_s = _winsorize_and_scale(ic_num, higher_is_better=True)
    return (de + ic_s) / 2.0

